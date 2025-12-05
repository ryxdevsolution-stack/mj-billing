import uuid
import io
import csv
from datetime import datetime
from flask import Blueprint, request, jsonify, g, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from extensions import db
from models.stock_model import StockEntry
from utils.auth_middleware import authenticate
from utils.permission_middleware import require_permission
from utils.audit_logger import log_action
from utils.cache_helper import get_cache_manager, invalidate_stock_cache

stock_bp = Blueprint('stock', __name__)

# Cache timeout for stock list (2 minutes - short since stock changes frequently)
STOCK_CACHE_TIMEOUT = 120


def generate_item_code(client_id, product_name):
    """
    Auto-generate item code based on product name and client
    Format: {PRODUCT_INITIALS}-{CLIENT_PREFIX}-{SEQUENCE}
    Example: LAP-550-001 for "Laptop" product
    """
    import re

    # Extract product initials (first 3 letters, alphanumeric only)
    clean_name = re.sub(r'[^a-zA-Z0-9]', '', product_name)
    product_prefix = clean_name[:3].upper() if clean_name else 'ITM'

    # Extract client prefix (first 3 chars of client_id)
    client_prefix = client_id[:3].upper()

    # Find the highest sequence number for this client
    existing_codes = StockEntry.query.filter_by(client_id=client_id).with_entities(StockEntry.item_code).all()

    max_sequence = 0
    for (code,) in existing_codes:
        if code:
            # Extract sequence number from codes matching pattern
            parts = code.split('-')
            if len(parts) >= 3:
                try:
                    seq = int(parts[-1])
                    max_sequence = max(max_sequence, seq)
                except ValueError:
                    continue

    # Generate next sequence
    next_sequence = max_sequence + 1

    # Format: PROD-CLI-001
    item_code = f"{product_prefix}-{client_prefix}-{next_sequence:03d}"

    # Handle collision (in case of race condition)
    counter = 0
    while StockEntry.query.filter_by(client_id=client_id, item_code=item_code).first():
        counter += 1
        next_sequence += counter
        item_code = f"{product_prefix}-{client_prefix}-{next_sequence:03d}"
        if counter > 100:  # Safety limit
            # Fallback to UUID-based code
            item_code = f"{product_prefix}-{client_prefix}-{uuid.uuid4().hex[:6].upper()}"
            break

    return item_code


@stock_bp.route('', methods=['POST'])
@authenticate
@require_permission('add_product')
def add_stock():
    """
    Add stock entry with client_id
    Auto-sum if product already exists for this client
    """
    try:
        data = request.get_json()
        client_id = g.user['client_id']

        # Validate required fields
        required_fields = ['product_name', 'quantity', 'rate']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        # Check if product already exists for this client
        existing_product = StockEntry.query.filter_by(
            client_id=client_id,
            product_name=data['product_name']
        ).first()

        if existing_product:
            # Auto-sum: Update existing product quantity
            old_data = existing_product.to_dict()
            existing_product.quantity += data['quantity']
            existing_product.rate = data.get('rate', existing_product.rate)
            existing_product.cost_price = data.get('cost_price', existing_product.cost_price)
            existing_product.mrp = data.get('mrp', existing_product.mrp)
            existing_product.pricing = data.get('pricing', existing_product.pricing)
            existing_product.category = data.get('category', existing_product.category)
            existing_product.unit = data.get('unit', existing_product.unit)
            existing_product.low_stock_alert = data.get('low_stock_alert', existing_product.low_stock_alert)

            # Handle item_code - auto-generate if empty/None
            item_code_value = data.get('item_code', existing_product.item_code)
            if isinstance(item_code_value, str):
                item_code_value = item_code_value.strip()
                item_code_value = item_code_value if item_code_value else None

            # If no item_code exists, auto-generate one
            if not item_code_value and not existing_product.item_code:
                item_code_value = generate_item_code(client_id, existing_product.product_name)

            existing_product.item_code = item_code_value

            # Handle barcode - set to None if empty to avoid unique constraint issues
            barcode_value = data.get('barcode', existing_product.barcode)
            if isinstance(barcode_value, str):
                barcode_value = barcode_value.strip()
                barcode_value = barcode_value if barcode_value else None
            existing_product.barcode = barcode_value

            existing_product.gst_percentage = data.get('gst_percentage', existing_product.gst_percentage)
            existing_product.hsn_code = data.get('hsn_code', existing_product.hsn_code)
            existing_product.updated_at = datetime.utcnow()

            db.session.commit()

            # Invalidate stock cache
            invalidate_stock_cache(client_id)

            # Log action
            log_action('UPDATE', 'stock_entry', existing_product.product_id, old_data, existing_product.to_dict())

            return jsonify({
                'success': True,
                'product_id': existing_product.product_id,
                'message': 'Stock updated successfully (quantity added)',
                'product': existing_product.to_dict()
            }), 200

        else:
            # Create new product entry
            # Handle barcode uniqueness - set to None if empty to avoid unique constraint issues
            barcode_value = data.get('barcode', '').strip()
            barcode_value = barcode_value if barcode_value else None

            # Handle item_code - auto-generate if empty
            item_code_value = data.get('item_code', '').strip()
            if not item_code_value:
                # Auto-generate item code
                item_code_value = generate_item_code(client_id, data['product_name'])
            # else: user provided item_code, use it

            new_product = StockEntry(
                product_id=str(uuid.uuid4()),
                client_id=client_id,
                product_name=data['product_name'],
                category=data.get('category', 'Other'),
                quantity=data['quantity'],
                rate=data['rate'],
                cost_price=data.get('cost_price'),
                mrp=data.get('mrp'),
                pricing=data.get('pricing'),
                unit=data.get('unit', 'pcs'),
                low_stock_alert=data.get('low_stock_alert', 10),
                item_code=item_code_value,
                barcode=barcode_value,
                gst_percentage=data.get('gst_percentage', 0),
                hsn_code=data.get('hsn_code', ''),
                created_at=datetime.utcnow()
            )

            db.session.add(new_product)
            db.session.commit()

            # Invalidate stock cache
            invalidate_stock_cache(client_id)

            # Log action
            log_action('CREATE', 'stock_entry', new_product.product_id, None, new_product.to_dict())

            return jsonify({
                'success': True,
                'product_id': new_product.product_id,
                'message': 'Stock added successfully',
                'product': new_product.to_dict()
            }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to add stock', 'message': str(e)}), 500


@stock_bp.route('', methods=['GET'])
@authenticate
@require_permission('view_stock')
def get_stock():
    """List stock entries filtered by client_id - OPTIMIZED with caching"""
    try:
        client_id = g.user['client_id']

        # Get query parameters
        category = request.args.get('category')
        search = request.args.get('search')
        limit = request.args.get('limit', type=int)  # Optional limit

        # Try cache for full list requests (no search/category filter)
        cache = get_cache_manager()
        if not category and not search and not limit:
            cache_key = f"stock:list:{client_id}"
            cached_data = cache.get(cache_key)
            if cached_data is not None:
                return jsonify({
                    'success': True,
                    'stock': cached_data
                }), 200

        # Build query
        query = StockEntry.query.filter_by(client_id=client_id)

        if category:
            query = query.filter_by(category=category)

        if search:
            query = query.filter(StockEntry.product_name.ilike(f'%{search}%'))

        # Apply ordering
        query = query.order_by(StockEntry.product_name)

        # Apply limit if provided (for better performance)
        if limit:
            query = query.limit(limit)

        # Get results
        stock_entries = query.all()
        stock_data = [entry.to_dict() for entry in stock_entries]

        # Cache full list for future requests
        if not category and not search and not limit:
            cache.set(f"stock:list:{client_id}", stock_data, STOCK_CACHE_TIMEOUT)

        return jsonify({
            'success': True,
            'stock': stock_data
        }), 200

    except Exception as e:
        return jsonify({'error': 'Failed to fetch stock', 'message': str(e)}), 500


@stock_bp.route('/alerts', methods=['GET'])
@authenticate
def get_low_stock_alerts():
    """Get low stock alerts filtered by client_id with full product details"""
    try:
        client_id = g.user['client_id']

        # Get products where quantity <= low_stock_alert
        low_stock = StockEntry.query.filter(
            StockEntry.client_id == client_id,
            StockEntry.quantity <= StockEntry.low_stock_alert
        ).order_by(StockEntry.quantity).all()

        return jsonify({
            'success': True,
            'alerts': [
                {
                    'product_id': item.product_id,
                    'product_name': item.product_name,
                    'current_quantity': item.quantity,
                    'alert_threshold': item.low_stock_alert,
                    'unit': item.unit
                }
                for item in low_stock
            ],
            'low_stock_products': [item.to_dict() for item in low_stock],
            'alert_count': len(low_stock)
        }), 200

    except Exception as e:
        return jsonify({'error': 'Failed to fetch alerts', 'message': str(e)}), 500


@stock_bp.route('/<product_id>', methods=['PUT'])
@authenticate
@require_permission('edit_product_details')
def update_stock(product_id):
    """Update stock entry with client_id validation"""
    try:
        client_id = g.user['client_id']
        data = request.get_json()

        # Find product
        product = StockEntry.query.filter_by(
            product_id=product_id,
            client_id=client_id
        ).first()

        if not product:
            return jsonify({'error': 'Product not found'}), 404

        # Store old data for audit
        old_data = product.to_dict()

        # Update fields
        if 'product_name' in data:
            product.product_name = data['product_name']
        if 'category' in data:
            product.category = data['category']
        if 'quantity' in data:
            product.quantity = data['quantity']
        if 'rate' in data:
            product.rate = data['rate']
        if 'cost_price' in data:
            product.cost_price = data['cost_price']
        if 'mrp' in data:
            product.mrp = data['mrp']
        if 'pricing' in data:
            product.pricing = data['pricing']
        if 'unit' in data:
            product.unit = data['unit']
        if 'low_stock_alert' in data:
            product.low_stock_alert = data['low_stock_alert']
        if 'item_code' in data:
            item_code_value = data['item_code'].strip() if isinstance(data['item_code'], str) else data['item_code']
            product.item_code = item_code_value if item_code_value else product.item_code

        # Auto-generate item_code if still missing
        if not product.item_code:
            product.item_code = generate_item_code(client_id, product.product_name)

        if 'barcode' in data:
            barcode_value = data['barcode'].strip() if isinstance(data['barcode'], str) else data['barcode']
            product.barcode = barcode_value if barcode_value else None
        if 'gst_percentage' in data:
            product.gst_percentage = data['gst_percentage']
        if 'hsn_code' in data:
            product.hsn_code = data['hsn_code']

        product.updated_at = datetime.utcnow()

        db.session.commit()

        # Log action
        log_action('UPDATE', 'stock_entry', product_id, old_data, product.to_dict())

        return jsonify({
            'success': True,
            'message': 'Stock updated successfully',
            'product': product.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to update stock', 'message': str(e)}), 500


@stock_bp.route('/<product_id>', methods=['DELETE'])
@authenticate
def delete_stock(product_id):
    """Delete stock entry with client_id validation"""
    try:
        client_id = g.user['client_id']

        # Find product
        product = StockEntry.query.filter_by(
            product_id=product_id,
            client_id=client_id
        ).first()

        if not product:
            return jsonify({'error': 'Product not found'}), 404

        # Store data for audit
        old_data = product.to_dict()

        db.session.delete(product)
        db.session.commit()

        # Log action
        log_action('DELETE', 'stock_entry', product_id, old_data, None)

        return jsonify({
            'success': True,
            'message': 'Stock deleted successfully'
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete stock', 'message': str(e)}), 500


@stock_bp.route('/bulk-import', methods=['POST'])
@authenticate
def bulk_import_stock():
    """
    Bulk import stock from CSV or Excel file
    Supports both create new and update existing products
    """
    try:
        client_id = g.user['client_id']

        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Check file extension
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        if file_ext not in ['csv', 'xlsx', 'xls']:
            return jsonify({'error': 'Invalid file format. Only CSV, XLSX, XLS files are allowed'}), 400

        # Read file based on extension
        try:
            if file_ext == 'csv':
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': 'Failed to read file', 'message': str(e)}), 400

        # Validate required columns
        required_columns = ['product_name', 'quantity', 'rate']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return jsonify({
                'error': f'Missing required columns: {", ".join(missing_columns)}',
                'required_columns': required_columns,
                'found_columns': list(df.columns)
            }), 400

        # Process each row
        success_count = 0
        error_count = 0
        updated_count = 0
        created_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # Skip rows with missing required fields
                if pd.isna(row['product_name']) or pd.isna(row['quantity']) or pd.isna(row['rate']):
                    error_count += 1
                    errors.append(f"Row {index + 2}: Missing required fields")
                    continue

                product_name = str(row['product_name']).strip()
                quantity = int(row['quantity'])
                rate = float(row['rate'])
                category = str(row['category']).strip() if 'category' in row and not pd.isna(row['category']) else 'Other'
                unit = str(row['unit']).strip() if 'unit' in row and not pd.isna(row['unit']) else 'pcs'
                low_stock_alert = int(row['low_stock_alert']) if 'low_stock_alert' in row and not pd.isna(row['low_stock_alert']) else 10
                cost_price = float(row['cost_price']) if 'cost_price' in row and not pd.isna(row['cost_price']) else None
                mrp = float(row['mrp']) if 'mrp' in row and not pd.isna(row['mrp']) else None
                pricing = float(row['pricing']) if 'pricing' in row and not pd.isna(row['pricing']) else None

                # Handle item_code - auto-generate if not provided
                item_code = str(row['item_code']).strip() if 'item_code' in row and not pd.isna(row['item_code']) else ''
                if not item_code:
                    item_code = generate_item_code(client_id, product_name)

                # Handle barcode - set to None if empty
                barcode = str(row['barcode']).strip() if 'barcode' in row and not pd.isna(row['barcode']) else ''
                barcode = barcode if barcode else None

                # Handle GST percentage and HSN code
                gst_percentage = float(row['gst_percentage']) if 'gst_percentage' in row and not pd.isna(row['gst_percentage']) else 0
                hsn_code = str(row['hsn_code']).strip() if 'hsn_code' in row and not pd.isna(row['hsn_code']) else ''

                # Validate quantity and rate
                if quantity < 0 or rate < 0:
                    error_count += 1
                    errors.append(f"Row {index + 2}: Quantity and rate must be positive")
                    continue

                # Check if product already exists
                existing_product = StockEntry.query.filter_by(
                    client_id=client_id,
                    product_name=product_name
                ).first()

                if existing_product:
                    # Update existing product (auto-sum quantity)
                    old_data = existing_product.to_dict()
                    existing_product.quantity += quantity
                    existing_product.rate = rate
                    existing_product.cost_price = cost_price if cost_price is not None else existing_product.cost_price
                    existing_product.mrp = mrp if mrp is not None else existing_product.mrp
                    existing_product.pricing = pricing if pricing is not None else existing_product.pricing
                    existing_product.category = category
                    existing_product.unit = unit
                    existing_product.low_stock_alert = low_stock_alert
                    existing_product.gst_percentage = gst_percentage
                    existing_product.hsn_code = hsn_code

                    # Auto-generate item_code if existing product doesn't have one
                    if not existing_product.item_code:
                        existing_product.item_code = item_code

                    # Update barcode if provided and existing doesn't have one
                    if barcode and not existing_product.barcode:
                        existing_product.barcode = barcode

                    existing_product.updated_at = datetime.utcnow()

                    # Log action
                    log_action('UPDATE', 'stock_entry', existing_product.product_id, old_data, existing_product.to_dict())

                    updated_count += 1
                else:
                    # Create new product
                    new_product = StockEntry(
                        product_id=str(uuid.uuid4()),
                        client_id=client_id,
                        product_name=product_name,
                        category=category,
                        quantity=quantity,
                        rate=rate,
                        cost_price=cost_price,
                        mrp=mrp,
                        pricing=pricing,
                        unit=unit,
                        low_stock_alert=low_stock_alert,
                        item_code=item_code,
                        barcode=barcode,
                        gst_percentage=gst_percentage,
                        hsn_code=hsn_code,
                        created_at=datetime.utcnow()
                    )

                    db.session.add(new_product)

                    # Log action
                    log_action('CREATE', 'stock_entry', new_product.product_id, None, new_product.to_dict())

                    created_count += 1

                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {index + 2}: {str(e)}")

        # Commit all changes
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Bulk import completed',
            'summary': {
                'total_rows': len(df),
                'success_count': success_count,
                'created_count': created_count,
                'updated_count': updated_count,
                'error_count': error_count,
                'errors': errors[:10] if errors else []  # Return first 10 errors
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to import stock', 'message': str(e)}), 500


@stock_bp.route('/download-template', methods=['POST'])
@authenticate
def download_template():
    """Download CSV/Excel template for bulk import with sample data"""
    try:
        import os
        data = request.get_json() or {}
        file_format = data.get('format', 'csv')

        print(file_format)

        # Get the path to template files
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')

        if file_format == 'csv':
            template_path = os.path.join(template_dir, 'stock_template.csv')
            if os.path.exists(template_path):
                return send_file(
                    template_path,
                    mimetype='text/csv',
                    as_attachment=True,
                    download_name='stock_import_template.csv'
                )
        else:  # xlsx
            template_path = os.path.join(template_dir, 'stock_template.xlsx')
            if os.path.exists(template_path):
                return send_file(
                    template_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='stock_import_template.xlsx'
                )

        # Fallback: Generate sample data if template files don't exist
        data = {
            'product_name': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Pen', 'Notebook'],
            'quantity': [10, 50, 30, 20, 100, 80],
            'rate': [50000.00, 500.00, 1500.00, 15000.00, 10.00, 25.00],
            'category': ['Electronics', 'Electronics', 'Electronics', 'Electronics', 'Stationery', 'Stationery'],
            'unit': ['pcs', 'pcs', 'pcs', 'pcs', 'pcs', 'pcs'],
            'low_stock_alert': [5, 10, 8, 5, 20, 15]
        }

        df = pd.DataFrame(data)
        output = io.BytesIO()

        if file_format == 'csv':
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name='stock_import_template.csv'
            )
        else:  # xlsx
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Stock Import')
            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='stock_import_template.xlsx'
            )

    except Exception as e:
        return jsonify({'error': 'Failed to generate template', 'message': str(e)}), 500


@stock_bp.route('/bulk-export', methods=['POST'])
@authenticate
def bulk_export_stock():
    """Export all stock data to CSV or Excel"""
    try:
        client_id = g.user['client_id']
        data = request.get_json() or {}
        file_format = data.get('format', 'csv')

        # Get all stock entries
        stock_entries = StockEntry.query.filter_by(client_id=client_id).order_by(StockEntry.product_name).all()

        if not stock_entries:
            return jsonify({'error': 'No stock data to export'}), 404

        # Convert to DataFrame
        data = []
        for stock in stock_entries:
            data.append({
                'product_name': stock.product_name,
                'quantity': stock.quantity,
                'rate': stock.rate,
                'category': stock.category,
                'unit': stock.unit,
                'low_stock_alert': stock.low_stock_alert,
                'created_at': stock.created_at.strftime('%Y-%m-%d %H:%M:%S') if stock.created_at else ''
            })

        df = pd.DataFrame(data)

        # Create in-memory file
        output = io.BytesIO()

        if file_format == 'csv':
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name='stock_export.csv'
            )
        else:  # xlsx
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Stock')
            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='stock_export.xlsx'
            )

    except Exception as e:
        return jsonify({'error': 'Failed to export stock', 'message': str(e)}), 500


@stock_bp.route('/lookup/<code>', methods=['GET'])
@authenticate
def lookup_product(code):
    """
    Quick product lookup by barcode or item code
    Returns product with rate, GST%, HSN, current stock
    Perfect for barcode scanner integration
    """
    try:
        client_id = g.user['client_id']

        # Search by barcode first (most common for scanner)
        product = StockEntry.query.filter_by(
            client_id=client_id,
            barcode=code
        ).first()

        # If not found by barcode, try item_code
        if not product:
            product = StockEntry.query.filter_by(
                client_id=client_id,
                item_code=code
            ).first()

        # If still not found, try exact product name match
        if not product:
            product = StockEntry.query.filter(
                StockEntry.client_id == client_id,
                StockEntry.product_name.ilike(code)
            ).first()

        if not product:
            return jsonify({
                'error': 'Product not found',
                'message': f'No product found with barcode, item code, or name: {code}'
            }), 404

        # Check stock availability
        stock_status = 'available'
        if product.quantity == 0:
            stock_status = 'out_of_stock'
        elif product.quantity <= product.low_stock_alert:
            stock_status = 'low_stock'

        return jsonify({
            'success': True,
            'product': product.to_dict(),
            'stock_status': stock_status,
            'available_quantity': product.quantity
        }), 200

    except Exception as e:
        return jsonify({'error': 'Failed to lookup product', 'message': str(e)}), 500


@stock_bp.route('/export-low-stock', methods=['POST'])
@authenticate
def export_low_stock():
    """Export low stock items to PDF or Excel for easy ordering"""
    try:
        client_id = g.user['client_id']
        data = request.get_json() or {}
        file_format = data.get('format', 'xlsx')

        # Get low stock items
        low_stock = StockEntry.query.filter(
            StockEntry.client_id == client_id,
            StockEntry.quantity <= StockEntry.low_stock_alert
        ).order_by(StockEntry.quantity).all()

        if not low_stock:
            return jsonify({'error': 'No low stock items to export'}), 404

        # Prepare data for export
        export_data = []
        total_cost = 0
        for item in low_stock:
            need_to_order = max(0, item.low_stock_alert - item.quantity)
            estimated_cost = need_to_order * float(item.rate)
            total_cost += estimated_cost

            export_data.append({
                'Product Name': item.product_name,
                'Category': item.category or '-',
                'Current Stock': f"{item.quantity} {item.unit}",
                'Min. Required': f"{item.low_stock_alert} {item.unit}",
                'Need to Order': f"{need_to_order} {item.unit}",
                'Rate per Unit': f"₹{float(item.rate):.2f}",
                'Estimated Cost': f"₹{estimated_cost:.2f}"
            })

        df = pd.DataFrame(export_data)

        # Add total row
        total_row = {
            'Product Name': 'TOTAL',
            'Category': '',
            'Current Stock': '',
            'Min. Required': '',
            'Need to Order': '',
            'Rate per Unit': '',
            'Estimated Cost': f"₹{total_cost:.2f}"
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        output = io.BytesIO()

        if file_format == 'pdf':
            # For PDF, we'll create a simple HTML and convert it
            # You can use libraries like reportlab or weasyprint for better PDF generation
            html_content = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #dc2626; }}
                    table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #fee; color: #991b1b; font-weight: bold; }}
                    .total-row {{ background-color: #f0f0f0; font-weight: bold; }}
                    .header {{ margin-bottom: 20px; }}
                    .alert {{ background-color: #fee; padding: 15px; border-left: 4px solid #dc2626; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>[WARNING] Low Stock Report - Order Required</h1>
                    <p><strong>Generated on:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p><strong>Total Items:</strong> {len(low_stock)}</p>
                </div>
                <div class="alert">
                    <p><strong>Action Required:</strong> These items need immediate restocking to maintain inventory levels.</p>
                </div>
                {df.to_html(index=False, classes='table', escape=False)}
                <p style="margin-top: 20px;"><em>This is an automatically generated report from your billing system.</em></p>
            </body>
            </html>
            """

            # Simple HTML to PDF conversion (basic approach)
            # For production, use proper PDF libraries
            output.write(html_content.encode('utf-8'))
            output.seek(0)

            return send_file(
                output,
                mimetype='text/html',  # Change to 'application/pdf' with proper PDF library
                as_attachment=True,
                download_name=f'low_stock_report_{datetime.now().strftime("%Y%m%d")}.html'
            )
        else:  # Excel
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Low Stock Report')

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Low Stock Report']

                # Style the header
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                header_font = Font(bold=True, color='991B1B')

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left')

                # Style the total row
                last_row = worksheet.max_row
                total_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
                total_font = Font(bold=True)

                for cell in worksheet[last_row]:
                    cell.fill = total_fill
                    cell.font = total_font

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'low_stock_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )

    except Exception as e:
        return jsonify({'error': 'Failed to export low stock report', 'message': str(e)}), 500
